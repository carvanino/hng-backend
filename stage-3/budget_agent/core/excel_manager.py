"""Google Drive Excel file management"""

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Dict, Optional
import logging
import io
import os
import tempfile

logger = logging.getLogger(__name__)


class ExcelManager:
    """Manages budget Excel files in user's Google Drive"""
    
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    
    def __init__(self, credentials: Credentials):
        """Initialize Drive service"""
        self.credentials = credentials
        self.service = build('drive', 'v3', credentials=credentials)
    
    def create_budget_file(self, user_id: str) -> str:
        """
        Create new budget tracker Excel file in user's Drive
        
        Returns:
            Google Drive file ID
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Get current month
            current_month = datetime.now().strftime('%B %Y')
            
            # Rename default sheet
            ws = wb.active
            ws.title = current_month
            
            # Set up the sheet
            self._setup_monthly_sheet(ws, current_month)
            
            # Save to temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name
            
            # Upload to Drive
            file_metadata = {
                'name': f'Budget Tracker {datetime.now().year}.xlsx',
                'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            media = MediaFileUpload(
                tmp_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                resumable=True
            )
            
            file = self.service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id, webViewLink'
            ).execute()
            
            # Clean up temp file
            os.unlink(tmp_path)
            
            file_id = file.get('id')
            logger.info(f"Created budget file in Drive: {file_id}")
            
            return file_id
            
        except Exception as e:
            logger.error(f"Error creating budget file: {e}")
            raise
    
    def _setup_monthly_sheet(self, ws, month_name: str):
        """Set up the structure of a monthly budget sheet"""
        
        # Title
        ws['A1'] = f'Zero Based Budget Template - {month_name}'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Income Section
        ws['A3'] = 'Income / Funds Received'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = 'Projected (₦)'
        ws['B3'].font = Font(bold=True)
        ws['C3'] = 'Actual (₦)'
        ws['C3'].font = Font(bold=True)
        ws['D3'] = 'Variance (₦)'
        ws['D3'].font = Font(bold=True)
        
        # Income rows (starting row 4)
        income_items = [
            'Salary',
            'Business Income',
            'Other Income'
        ]
        
        for idx, item in enumerate(income_items, start=4):
            ws[f'A{idx}'] = item
        
        total_income_row = 4 + len(income_items)
        ws[f'A{total_income_row}'] = 'Total Income'
        ws[f'A{total_income_row}'].font = Font(bold=True)
        
        # Expenses Section Header
        expense_header_row = total_income_row + 2
        ws[f'A{expense_header_row}'] = 'Expenses'
        ws[f'A{expense_header_row}'].font = Font(bold=True)
        
        # Daily Expenditure Summary Header
        daily_start_row = expense_header_row + 2
        ws[f'A{daily_start_row}'] = 'Day'
        ws[f'B{daily_start_row}'] = 'Expenses (₦)'
        ws[f'C{daily_start_row}'] = 'Income (₦)'
        ws[f'D{daily_start_row}'] = 'Category'
        ws[f'E{daily_start_row}'] = 'Description'
        
        # Make headers bold
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{daily_start_row}'].font = Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 40
    
    def add_transaction(self, file_id: str, transaction: Dict, category: str) -> bool:
        """
        Add transaction to the Excel file
        
        Args:
            file_id: Google Drive file ID
            transaction: Transaction dict with amount, date, description, type
            category: Category for the transaction
        """
        try:
            # Download file from Drive
            wb = self._download_workbook(file_id)
            
            # Get or create sheet for current month
            month_name = datetime.now().strftime('%B %Y')
            if month_name not in wb.sheetnames:
                ws = wb.create_sheet(month_name)
                self._setup_monthly_sheet(ws, month_name)
            else:
                ws = wb[month_name]
            
            # Find next empty row in daily expenditure section
            next_row = self._find_next_row(ws)
            
            # Extract day from transaction date
            try:
                trans_date = datetime.strptime(transaction['date'], '%Y-%m-%d')
                day = trans_date.day
            except:
                day = datetime.now().day
            
            # Add transaction
            ws[f'A{next_row}'] = day
            
            if transaction['type'] == 'expense':
                ws[f'B{next_row}'] = transaction['amount']
                ws[f'C{next_row}'] = None
            else:  # income
                ws[f'B{next_row}'] = None
                ws[f'C{next_row}'] = transaction['amount']
            
            ws[f'D{next_row}'] = category
            ws[f'E{next_row}'] = transaction['description']
            
            # Upload back to Drive
            self._upload_workbook(file_id, wb)
            
            logger.info(f"Added transaction to file {file_id}: {category} - ₦{transaction['amount']}")
            return True
            
        except Exception as e:
            logger.error(f"Error adding transaction: {e}")
            return False
    
    def _find_next_row(self, ws) -> int:
        """Find next empty row in daily expenditure section"""
        # Daily expenditure starts around row 11, but let's search for it
        # Find the "Day" header
        for row in range(1, 20):
            if ws[f'A{row}'].value == 'Day':
                # Start from next row and find first empty
                check_row = row + 1
                while ws[f'A{check_row}'].value is not None or ws[f'B{check_row}'].value is not None:
                    check_row += 1
                return check_row
        
        # Fallback
        return 15
    
    def _download_workbook(self, file_id: str) -> Workbook:
        """Download Excel file from Drive and return as Workbook"""
        request = self.service.files().get_media(fileId=file_id)
        
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        fh.seek(0)
        return load_workbook(fh)
    
    def _upload_workbook(self, file_id: str, wb: Workbook):
        """Upload Workbook to Drive"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        media = MediaFileUpload(
            tmp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        self.service.files().update(
            fileId=file_id,
            media_body=media
        ).execute()
        
        os.unlink(tmp_path)
    
    def get_file_url(self, file_id: str) -> Optional[str]:
        """Get web view URL for file"""
        try:
            file = self.service.files().get(
                fileId=file_id,
                fields='webViewLink'
            ).execute()
            return file.get('webViewLink')
        except HttpError as e:
            logger.error(f"Error getting file URL: {e}")
            return None
